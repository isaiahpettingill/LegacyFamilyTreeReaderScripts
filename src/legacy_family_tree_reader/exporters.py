"""GEDCOM and spreadsheet exports for the merged SQLite database.

The merged database deliberately uses descriptive names, but older merged files
used a few singular table and column names.  The small amount of schema
introspection here keeps exports useful for those files without coupling them to
column order or to Legacy's original ``tblXX`` schema.
"""

from __future__ import annotations

import re
import sqlite3
from collections import defaultdict
from collections.abc import Iterable, Mapping, Sequence
from pathlib import Path
from typing import Any, TypeAlias

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_Row: TypeAlias = dict[str, Any]
_Key: TypeAlias = tuple[str, str]

_TABLES: Mapping[str, tuple[str, ...]] = {
    "datasets": ("data_sets", "datasets", "dataset"),
    "people": ("individuals", "people", "persons", "person", "individual"),
    "alternate_names": ("alternate_names", "aliases", "other_names"),
    "families": ("marriages", "families", "family", "marriage_records"),
    "children": (
        "parent_child",
        "parent_children",
        "family_children",
        "children",
        "child_links",
    ),
    "events": ("events", "event_records", "event"),
    "event_participants": ("event_participants", "event_participant", "participants"),
    "event_types": ("event_types", "event_type"),
    "sources": ("sources", "source_records", "source"),
    "source_types": ("source_types", "source_type"),
    "citations": ("citations", "source_citations", "citation"),
    "media": ("media", "multimedia", "media_records"),
    "media_paths": ("media_paths", "multimedia_paths", "media_path"),
    "locations": ("locations", "places", "location", "place"),
    "marriage_statuses": ("marriage_statuses", "family_statuses", "marriage_status"),
    "child_statuses": ("child_statuses", "child_status"),
    "child_relationships": (
        "child_relationship_types",
        "child_relationships",
        "parent_relationships",
    ),
}

_VITAL_FACTS = (
    (
        "Birth",
        "BIRT",
        ("birth_date", "birth_d"),
        ("birth_location_id", "birth_place_id", "location_id_birth"),
        ("birth_location", "birth_place"),
        ("birth_notes", "birth_note"),
    ),
    (
        "Christening",
        "CHR",
        ("christening_date", "christening_d", "christening", "christ_date", "chris_d"),
        ("christening_location_id", "christening_place_id", "location_id_christening"),
        ("christening_location", "christening_place"),
        ("christening_notes", "christening_note", "chris_note"),
    ),
    (
        "Death",
        "DEAT",
        ("death_date", "death_d"),
        ("death_location_id", "death_place_id", "location_id_death"),
        ("death_location", "death_place"),
        ("death_notes", "death_note"),
    ),
    (
        "Burial",
        "BURI",
        ("burial_date", "buried_date", "buried_d"),
        ("burial_location_id", "burial_place_id", "location_id_burial"),
        ("burial_location", "burial_place"),
        ("burial_notes", "buried_note", "burial_note"),
    ),
)

_EVENT_TAGS = {
    "adoption": "ADOP",
    "baptism": "BAPM",
    "bar mitzvah": "BARM",
    "bat mitzvah": "BASM",
    "blessing": "BLES",
    "census": "CENS",
    "confirmation": "CONF",
    "cremation": "CREM",
    "education": "EDUC",
    "emigration": "EMIG",
    "first communion": "FCOM",
    "graduation": "GRAD",
    "immigration": "IMMI",
    "military service": "MILI",
    "naturalization": "NATU",
    "occupation": "OCCU",
    "ordination": "ORDN",
    "probate": "PROB",
    "residence": "RESI",
    "retirement": "RETI",
    "will": "WILL",
}

_STANDARD_EVENT_TAGS = set(_EVENT_TAGS.values()) | {
    "ADOP",
    "BIRT",
    "BURI",
    "CHR",
    "DEAT",
    "DIV",
    "DIVF",
    "ENGA",
    "EVEN",
    "MARB",
    "MARC",
    "MARL",
    "MARR",
    "MARS",
}


def _normalise(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", name.casefold())


def _quote_identifier(name: str) -> str:
    return f'"{name.replace(chr(34), chr(34) * 2)}"'


class _Schema:
    def __init__(self, connection: sqlite3.Connection) -> None:
        rows = connection.execute(
            "SELECT name FROM sqlite_master WHERE type IN ('table', 'view')"
        ).fetchall()
        self._tables = {_normalise(row[0]): row[0] for row in rows}
        self._columns: dict[str, dict[str, str]] = {}
        self._connection = connection

    def table(self, logical_name: str) -> str | None:
        for candidate in _TABLES[logical_name]:
            if table := self._tables.get(_normalise(candidate)):
                return table
        return None

    def columns(self, table: str) -> dict[str, str]:
        if table not in self._columns:
            quoted = _quote_identifier(table)
            rows = self._connection.execute(f"PRAGMA table_info({quoted})").fetchall()
            self._columns[table] = {_normalise(row[1]): row[1] for row in rows}
        return self._columns[table]

    def rows(self, logical_name: str, dataset_id: int | str | None) -> list[_Row]:
        table = self.table(logical_name)
        if table is None:
            return []

        columns = self.columns(table)
        dataset_column = columns.get("datasetid")
        if logical_name == "datasets":
            dataset_column = columns.get("id", dataset_column)
        sql = f"SELECT * FROM {_quote_identifier(table)}"
        parameters: tuple[Any, ...] = ()
        if dataset_id is not None and dataset_column is not None:
            sql += f" WHERE {_quote_identifier(dataset_column)} = ?"
            parameters = (dataset_id,)
        cursor = self._connection.execute(sql, parameters)
        names = [description[0] for description in cursor.description or ()]
        return [
            {_normalise(name): value for name, value in zip(names, values, strict=True)}
            for values in cursor.fetchall()
        ]


def _connect_read_only(db_path: str | Path) -> sqlite3.Connection:
    path = Path(db_path).expanduser().resolve(strict=True)
    connection = sqlite3.connect(f"{path.as_uri()}?mode=ro", uri=True)
    connection.execute("PRAGMA query_only = ON")
    return connection


def _get(row: Mapping[str, Any], *names: str, default: Any = None) -> Any:
    for name in names:
        value = row.get(_normalise(name))
        if value is not None:
            return value
    return default


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def _present(value: Any) -> bool:
    return value is not None and _text(value).strip() != ""


def _has_id(value: Any) -> bool:
    """Return whether a Legacy foreign key refers to a real record."""
    return _present(value) and _id(value) != "0"


def _truthy(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip().casefold() in {"1", "true", "yes", "y", "private", "living"}
    return bool(value)


def _id(value: Any) -> str:
    return _text(value).strip()


def _dataset(row: Mapping[str, Any], requested: int | str | None) -> str:
    value = _get(row, "dataset_id")
    if value is None:
        value = requested if requested is not None else 0
    return _id(value)


def _key(dataset: Any, record_id: Any) -> _Key:
    return (_id(dataset), _id(record_id))


def _record_id(row: Mapping[str, Any], kind: str) -> Any:
    aliases: Mapping[str, tuple[str, ...]] = {
        "person": ("person_id", "individual_id", "id"),
        "family": ("family_id", "marriage_id", "id"),
        "event": ("event_id", "id"),
        "source": ("source_id", "id"),
        "citation": ("citation_id", "id"),
        "media": ("media_id", "multimedia_id", "id"),
        "location": ("location_id", "place_id", "id"),
        "event_type": ("event_type_id", "type_id", "id"),
        "dataset": ("dataset_id", "id"),
    }
    return _get(row, *aliases[kind])


def _sort_rows(rows: Iterable[_Row], requested: int | str | None, kind: str) -> list[_Row]:
    return sorted(
        rows,
        key=lambda row: (
            _dataset(row, requested).casefold(),
            _id(_record_id(row, kind)).casefold(),
        ),
    )


def _xref_part(value: Any) -> str:
    raw = _id(value)
    if not raw:
        return "0"
    parts: list[str] = []
    for byte in raw.encode("utf-8"):
        char = chr(byte)
        parts.append(char if char.isascii() and char.isalnum() else f"_{byte:02X}")
    return "".join(parts)


def _xref(prefix: str, dataset: Any, record_id: Any) -> str:
    return f"@{prefix}D{_xref_part(dataset)}R{_xref_part(record_id)}@"


def _full_name(row: Mapping[str, Any]) -> str:
    explicit = _get(row, "full_name", "display_name", "name")
    if _present(explicit):
        return _text(explicit).strip()
    parts = (
        _get(row, "prefix", "name_prefix", "title_prefix"),
        _get(row, "given_name", "given_names", "given", "first_name"),
        _get(row, "middle_name", "middle"),
        _get(row, "surname", "last_name", "family_name"),
        _get(row, "suffix", "name_suffix", "title", "title_suffix"),
    )
    return " ".join(_text(part).strip() for part in parts if _present(part))


def _gedcom_name(row: Mapping[str, Any]) -> str:
    given_parts = (
        _get(row, "given_name", "given_names", "given", "first_name"),
        _get(row, "middle_name", "middle"),
    )
    given = " ".join(_text(part).strip() for part in given_parts if _present(part))
    surname = _text(_get(row, "surname", "last_name", "family_name")).strip()
    given = given.replace("/", " ")
    surname = surname.replace("/", " ")
    if given or surname:
        return f"{given} /{surname}/".strip()
    return f"{_full_name(row).replace('/', ' ')} //".strip()


def _sex(value: Any, *, legacy_gender: bool = False) -> str:
    text = _text(value).strip().casefold()
    if legacy_gender:
        return {"0": "M", "1": "F", "2": "U"}.get(text, "U")
    if text in {"1", "m", "male", "man"}:
        return "M"
    if text in {"2", "f", "female", "woman"}:
        return "F"
    return "U"


def _person_sex(row: Mapping[str, Any]) -> str:
    gender = _get(row, "gender", "gender_code")
    if gender is not None:
        return _sex(gender, legacy_gender=True)
    return _sex(_get(row, "sex"))


_MONTHS = {
    "jan": "JAN",
    "january": "JAN",
    "feb": "FEB",
    "february": "FEB",
    "mar": "MAR",
    "march": "MAR",
    "apr": "APR",
    "april": "APR",
    "may": "MAY",
    "jun": "JUN",
    "june": "JUN",
    "jul": "JUL",
    "july": "JUL",
    "aug": "AUG",
    "august": "AUG",
    "sep": "SEP",
    "sept": "SEP",
    "september": "SEP",
    "oct": "OCT",
    "october": "OCT",
    "nov": "NOV",
    "november": "NOV",
    "dec": "DEC",
    "december": "DEC",
}
_MONTH_NUMBERS = (
    "JAN",
    "FEB",
    "MAR",
    "APR",
    "MAY",
    "JUN",
    "JUL",
    "AUG",
    "SEP",
    "OCT",
    "NOV",
    "DEC",
)


def _atomic_gedcom_date(value: str) -> str | None:
    value = value.strip(" .,")
    iso = re.fullmatch(r"(\d{4})[-/](\d{1,2})(?:[-/](\d{1,2}))?", value)
    if iso:
        year, month, day = iso.groups()
        if not 1 <= int(month) <= 12 or day is not None and not 1 <= int(day) <= 31:
            return None
        month_name = _MONTH_NUMBERS[int(month) - 1]
        return f"{int(day)} {month_name} {year}" if day else f"{month_name} {year}"

    year_only = re.fullmatch(r"\d{3,4}", value)
    if year_only:
        return value

    day_month_year = re.fullmatch(r"(\d{1,2})[\s./-]+([A-Za-z]+)[\s,./-]+(\d{3,4})", value)
    if day_month_year:
        day, month, year = day_month_year.groups()
        normalised_month = _MONTHS.get(month.casefold())
        if normalised_month and 1 <= int(day) <= 31:
            return f"{int(day)} {normalised_month} {year}"
        return None

    month_day_year = re.fullmatch(r"([A-Za-z]+)\s+(\d{1,2}),?\s+(\d{3,4})", value)
    if month_day_year:
        month, day, year = month_day_year.groups()
        normalised_month = _MONTHS.get(month.casefold())
        if normalised_month and 1 <= int(day) <= 31:
            return f"{int(day)} {normalised_month} {year}"
        return None

    month_year = re.fullmatch(r"([A-Za-z]+)[\s,./-]+(\d{3,4})", value)
    if month_year and (month := _MONTHS.get(month_year.group(1).casefold())):
        return f"{month} {month_year.group(2)}"
    return None


def _gedcom_date(value: Any) -> str | None:
    """Convert only unambiguous Legacy display dates to GEDCOM 5.5.1 dates."""
    raw = _text(value).replace("\x00", " ").strip()
    if not raw:
        return None
    raw = re.sub(r"\s+", " ", raw)

    packed = re.fullmatch(
        r"(?P<qualifier>\d{2})(?P<day>\d{2})(?P<month>\d{2})"
        r"(?P<year>\d{4})(?P<tail>\d{8})",
        raw,
    )
    if packed:
        if packed["qualifier"] != "00" or packed["tail"] != "00000000":
            return None
        day = int(packed["day"])
        month = int(packed["month"])
        year = packed["year"]
        if day == 0 and month == 0 and int(year) > 0:
            return year
        if day == 0 and 1 <= month <= 12 and int(year) > 0:
            return f"{_MONTH_NUMBERS[month - 1]} {year}"
        if 1 <= day <= 31 and 1 <= month <= 12 and int(year) > 0:
            return f"{day} {_MONTH_NUMBERS[month - 1]} {year}"
        return None

    qualifiers = (
        (r"^(?:abt|about|ca|circa|c)\.?\s+", "ABT"),
        (r"^(?:bef|before)\.?\s+|^<\s*", "BEF"),
        (r"^(?:aft|after)\.?\s+|^>\s*", "AFT"),
    )
    for pattern, gedcom_qualifier in qualifiers:
        if match := re.match(pattern, raw, flags=re.IGNORECASE):
            date = _atomic_gedcom_date(raw[match.end() :])
            return f"{gedcom_qualifier} {date}" if date else None

    between = re.fullmatch(
        r"(?:bet(?:ween)?\.?\s+)(.+?)\s+(?:and|&)\s+(.+)", raw, flags=re.IGNORECASE
    )
    if between:
        first = _atomic_gedcom_date(between.group(1))
        second = _atomic_gedcom_date(between.group(2))
        return f"BET {first} AND {second}" if first and second else None

    date_range = re.fullmatch(r"from\s+(.+?)\s+to\s+(.+)", raw, flags=re.IGNORECASE)
    if date_range:
        first = _atomic_gedcom_date(date_range.group(1))
        second = _atomic_gedcom_date(date_range.group(2))
        return f"FROM {first} TO {second}" if first and second else None

    return _atomic_gedcom_date(raw)


def _clean_gedcom_text(value: Any) -> str:
    text = _text(value).replace("\r\n", "\n").replace("\r", "\n")
    text = "".join(char if char == "\n" or ord(char) >= 32 else " " for char in text)
    return text.replace("@", "@@")


def _take_utf8(text: str, limit: int) -> tuple[str, str]:
    if len(text.encode("utf-8")) <= limit:
        return text, ""
    size = 0
    end = 0
    for end, char in enumerate(text, start=1):
        encoded_size = len(char.encode("utf-8"))
        if size + encoded_size > limit:
            end -= 1
            break
        size += encoded_size
    return text[:end], text[end:]


def _wrapped_lines(level: int, tag: str, value: Any, *, max_bytes: int = 248) -> list[str]:
    text = _clean_gedcom_text(value)
    if not text:
        return []
    result: list[str] = []
    logical_lines = text.split("\n")
    for line_number, logical_line in enumerate(logical_lines):
        current_tag = tag if line_number == 0 else "CONT"
        current_level = level if line_number == 0 else level + 1
        remainder = logical_line
        first_chunk = True
        while remainder or first_chunk:
            prefix = f"{current_level} {current_tag}"
            available = max(1, max_bytes - len(prefix.encode("ascii")) - 1)
            chunk, remainder = _take_utf8(remainder, available)
            result.append(f"{prefix} {chunk}".rstrip())
            first_chunk = False
            current_tag = "CONC"
            current_level = level + 1
    return result


def _location_name(row: Mapping[str, Any]) -> str:
    return _text(
        _get(row, "location", "place", "name", "full_name", "sorted_location", "short_name")
    ).strip()


def _media_path_lookup(rows: Iterable[_Row], requested: int | str | None) -> dict[_Key, _Row]:
    result: dict[_Key, _Row] = {}
    for row in rows:
        path_id = _get(row, "media_path_id", "path_id", "id")
        if _has_id(path_id):
            result[_key(_dataset(row, requested), path_id)] = row
    return result


def _media_file(row: Mapping[str, Any], dataset: Any, media_paths: Mapping[_Key, _Row]) -> str:
    name = ""
    for field in (
        "path",
        "file_path",
        "media_name_url",
        "filename",
        "file_name",
        "media_name",
        "url",
    ):
        value = row.get(_normalise(field))
        if _present(value):
            name = _text(value).strip()
            break
    if not name or name.startswith(("/", "\\")) or re.match(r"^[A-Za-z]:[\\/]", name):
        return name
    path_id = _get(row, "picture_media_path_id", "media_path_id", "path_id")
    path_row = _resolve_row(media_paths, dataset, path_id)
    base = _text(_get(path_row or {}, "media_path", "path", "base_path")).strip()
    if not base:
        return name
    base = base.rstrip("/\\")
    name = name.lstrip("/\\")
    return f"{base}/{name}"


def _lookup(rows: Iterable[_Row], requested: int | str | None, kind: str) -> dict[_Key, _Row]:
    result: dict[_Key, _Row] = {}
    for row in rows:
        record_id = _record_id(row, kind)
        if record_id is not None:
            result[_key(_dataset(row, requested), record_id)] = row
    return result


def _column_lookup(
    rows: Iterable[_Row], requested: int | str | None, *id_names: str
) -> dict[_Key, _Row]:
    result: dict[_Key, _Row] = {}
    for row in rows:
        record_id = _get(row, *id_names)
        if _has_id(record_id):
            result[_key(_dataset(row, requested), record_id)] = row
    return result


def _resolve_row(lookup: Mapping[_Key, _Row], dataset: Any, record_id: Any) -> _Row | None:
    if not _has_id(record_id):
        return None
    exact = lookup.get(_key(dataset, record_id))
    if exact is not None:
        return exact
    matches = [row for (row_dataset, row_id), row in lookup.items() if row_id == _id(record_id)]
    return matches[0] if len(matches) == 1 else None


def _place(
    row: Mapping[str, Any],
    dataset: Any,
    location_lookup: Mapping[_Key, _Row],
    id_names: Sequence[str],
    text_names: Sequence[str],
) -> str:
    explicit = _get(row, *text_names)
    if _present(explicit):
        return _text(explicit).strip()
    location_id = _get(row, *id_names)
    location = _resolve_row(location_lookup, dataset, location_id)
    return _location_name(location) if location else ""


def _private_person(row: Mapping[str, Any]) -> bool:
    return _truthy(_get(row, "private", "is_private", "private_flag")) or _truthy(
        _get(row, "living", "is_living", "living_flag")
    )


def _owner_kind(value: Any) -> str:
    text = _text(value).strip().casefold().replace("_", " ")
    if text in {"i", "indi", "individual", "person", "people", "0"}:
        return "person"
    if text in {"f", "fam", "family", "marriage", "1"}:
        return "family"
    if text in {"e", "event", "2"}:
        return "event"
    if text in {"s", "source", "3"}:
        return "source"
    return text


def _numeric_code(value: Any) -> int | None:
    text = _text(value).strip()
    return int(text) if re.fullmatch(r"-?\d+", text) else None


def _owner(row: Mapping[str, Any]) -> tuple[str, Any]:
    export_owner_id = _get(row, "export_owner_id")
    if _has_id(export_owner_id):
        return _text(_get(row, "export_owner_kind")).strip(), export_owner_id
    kind_value = _get(row, "owner_type", "record_type", "entity_type", "target_type")
    owner_id = _get(
        row,
        "owner_id",
        "owner_record_id",
        "record_id",
        "entity_id",
        "target_id",
        "subject_id",
    )
    if _present(kind_value):
        if not _has_id(owner_id):
            owner_id = _get(row, "individual_id")
        if _has_id(owner_id):
            return _owner_kind(kind_value), owner_id
    if _has_id(person_id := _get(row, "person_id")):
        return "person", person_id
    if _has_id(family_id := _get(row, "family_id", "marriage_id")):
        return "family", family_id
    if _has_id(individual_id := _get(row, "individual_id")):
        return "person", individual_id
    if _has_id(event_id := _get(row, "event_id")):
        return "event", event_id
    return (
        _owner_kind(kind_value),
        owner_id,
    )


def _event_owner(
    row: Mapping[str, Any],
    dataset: str,
    people_by_legacy_id: Mapping[_Key, _Row],
    families_by_legacy_id: Mapping[_Key, _Row],
    people_by_internal_id: Mapping[_Key, _Row],
    families_by_internal_id: Mapping[_Key, _Row],
) -> tuple[str, Any]:
    code = _numeric_code(_get(row, "record_type"))
    owner_id = _get(row, "owner_record_id")
    if code in {0, 20} and _has_id(owner_id):
        kind = "person" if code == 0 else "family"
        legacy_lookup = people_by_legacy_id if code == 0 else families_by_legacy_id
        internal_lookup = people_by_internal_id if code == 0 else families_by_internal_id
        owner = _resolve_row(legacy_lookup, dataset, owner_id)
        if owner is None:
            owner = _resolve_row(internal_lookup, dataset, owner_id)
        if owner is not None:
            return kind, _record_id(owner, "person" if code == 0 else "family")
        return kind, owner_id
    return _owner(row)


def _normalise_event_owners(
    events: Iterable[_Row],
    requested: int | str | None,
    people_by_legacy_id: Mapping[_Key, _Row],
    families_by_legacy_id: Mapping[_Key, _Row],
    people_by_internal_id: Mapping[_Key, _Row],
    families_by_internal_id: Mapping[_Key, _Row],
) -> list[_Row]:
    result: list[_Row] = []
    for event in events:
        normalised = dict(event)
        kind, owner_id = _event_owner(
            event,
            _dataset(event, requested),
            people_by_legacy_id,
            families_by_legacy_id,
            people_by_internal_id,
            families_by_internal_id,
        )
        normalised["exportownerkind"] = kind
        normalised["exportownerid"] = owner_id
        result.append(normalised)
    return result


def _media_owner(row: Mapping[str, Any]) -> tuple[str, Any]:
    owner_id = _get(row, "owner_record_id")
    code = _numeric_code(_get(row, "record_type"))
    if code is not None and _has_id(owner_id):
        kinds = {
            0: "person",
            1: "birth",
            2: "christening",
            3: "death",
            4: "burial",
            20: "family",
            30: "event",
            40: "source",
            41: "citation",
            70: "todo",
        }
        return kinds.get(code, f"record type {code}"), owner_id
    return _owner(row)


def _event_type(
    row: Mapping[str, Any], dataset: Any, event_type_lookup: Mapping[_Key, _Row]
) -> tuple[str, str]:
    type_row = _resolve_row(event_type_lookup, dataset, _get(row, "event_type_id", "type_id"))
    name = _get(row, "event_type", "event_name", "type_name")
    if not _present(name) and type_row:
        name = _get(type_row, "event_type", "event_name", "type_name", "name")
    name_text = _text(name).strip() or "Event"
    supplied_tag = _get(row, "gedcom_tag", "ged_tag", "tag")
    if not _present(supplied_tag) and type_row:
        supplied_tag = _get(type_row, "gedcom_tag", "ged_tag", "tag")
    tag = _text(supplied_tag).strip().upper()
    if tag not in _STANDARD_EVENT_TAGS:
        tag = _EVENT_TAGS.get(name_text.casefold(), "EVEN")
    return name_text, tag


def _citation_targets(
    row: Mapping[str, Any],
    alternate_name_lookup: Mapping[_Key, _Row],
    child_lookup: Mapping[_Key, _Row],
    event_participant_lookup: Mapping[_Key, _Row],
    requested: int | str | None,
) -> tuple[tuple[str, Any], ...]:
    if _present(_get(row, "owner_type", "entity_type", "target_type")):
        kind, owner_id = _owner(row)
        return ((kind, owner_id),) if _has_id(owner_id) else ()

    cited_id = _get(row, "cited_record_id")
    code = _numeric_code(_get(row, "type", "citation_type"))
    if code is None or not _has_id(cited_id):
        return ()
    direct_kinds = {
        0: "person",
        1: "person",
        2: "birth",
        3: "christening",
        4: "death",
        5: "burial",
        12: "todo description",
        13: "todo results",
        15: "person",
        16: "person",
        18: "family",
        20: "family",
        26: "person",
        27: "person",
        28: "story",
        30: "event",
    }
    if code in direct_kinds:
        return ((direct_kinds[code], cited_id),)

    dataset = _dataset(row, requested)
    if code == 10:
        alternate_name = _resolve_row(alternate_name_lookup, dataset, cited_id)
        person_id = _get(alternate_name or {}, "individual_id", "person_id")
        return (("person", person_id),) if _has_id(person_id) else (("alternate name", cited_id),)
    if code == 17:
        child = _resolve_row(child_lookup, dataset, cited_id)
        if child is None:
            return (("child", cited_id),)
        family_id, person_id = _child_link(child)
        return tuple(
            (kind, owner_id)
            for kind, owner_id in (("person", person_id), ("family", family_id))
            if _has_id(owner_id)
        )
    if code == 31:
        participant = _resolve_row(event_participant_lookup, dataset, cited_id)
        event_id = _get(participant or {}, "event_id")
        return (("event", event_id),) if _has_id(event_id) else (("event participant", cited_id),)
    return ((f"citation type {code}", cited_id),)


def _citation_target(
    row: Mapping[str, Any],
    alternate_name_lookup: Mapping[_Key, _Row],
    child_lookup: Mapping[_Key, _Row],
    event_participant_lookup: Mapping[_Key, _Row],
    requested: int | str | None,
) -> tuple[str, Any]:
    targets = _citation_targets(
        row,
        alternate_name_lookup,
        child_lookup,
        event_participant_lookup,
        requested,
    )
    return targets[0] if targets else ("", None)


def _citation_lines(citation: Mapping[str, Any], source_xref: str) -> list[str]:
    lines = [f"1 SOUR {source_xref}"]
    detail = _get(citation, "detail", "source_detail", "page", "where_within_source")
    if _present(detail):
        lines.extend(_wrapped_lines(2, "PAGE", detail))
    text = _get(
        citation,
        "text",
        "source_text",
        "detail_text",
        "source_detail_text",
        "content",
        "quote",
    )
    if _present(text):
        lines.append("2 DATA")
        lines.extend(_wrapped_lines(3, "TEXT", text))
    notes = _get(citation, "notes", "note", "detail_note", "source_detail_note")
    if _present(notes):
        lines.extend(_wrapped_lines(2, "NOTE", notes))
    quality = _get(citation, "quality", "surety", "source_surety", "certainty")
    if _present(quality):
        quality_text = _text(quality).strip()
        quality_map = {
            "unreliable": "0",
            "questionable": "1",
            "secondary": "2",
            "direct": "3",
            "primary": "3",
        }
        if quality_text in {"0", "1", "2", "3"}:
            lines.append(f"2 QUAY {quality_text}")
        elif quality_text.casefold() in quality_map:
            lines.append(f"2 QUAY {quality_map[quality_text.casefold()]}")
    return lines


def _citation_map(
    citations: Iterable[_Row],
    source_lookup: Mapping[_Key, _Row],
    alternate_name_lookup: Mapping[_Key, _Row],
    child_lookup: Mapping[_Key, _Row],
    event_participant_lookup: Mapping[_Key, _Row],
    requested: int | str | None,
) -> dict[tuple[str, str, str], list[tuple[_Row, str]]]:
    result: dict[tuple[str, str, str], list[tuple[_Row, str]]] = defaultdict(list)
    for citation in citations:
        dataset = _dataset(citation, requested)
        source_id = _get(citation, "source_id")
        source = _resolve_row(source_lookup, dataset, source_id)
        if source is None:
            continue
        source_record_id = _record_id(source, "source")
        if not _has_id(source_record_id):
            continue
        source_xref = _xref("S", _dataset(source, requested), source_record_id)
        for owner_kind, owner_id in _citation_targets(
            citation,
            alternate_name_lookup,
            child_lookup,
            event_participant_lookup,
            requested,
        ):
            if owner_kind and _has_id(owner_id):
                result[(dataset, owner_kind, _id(owner_id))].append((citation, source_xref))
    return result


def _append_citations(
    lines: list[str],
    citation_map: Mapping[tuple[str, str, str], list[tuple[_Row, str]]],
    dataset: str,
    kinds: Iterable[str],
    owner_id: Any,
    *,
    level: int = 1,
) -> None:
    seen: set[tuple[str, str]] = set()
    for kind in kinds:
        for citation, source_xref in citation_map.get((dataset, kind, _id(owner_id)), []):
            citation_id = _id(_record_id(citation, "citation")) or repr(sorted(citation.items()))
            marker = (citation_id, source_xref)
            if marker not in seen:
                for line in _citation_lines(citation, source_xref):
                    current_level, remainder = line.split(" ", 1)
                    lines.append(f"{int(current_level) + level - 1} {remainder}")
                seen.add(marker)


def _append_date(lines: list[str], level: int, raw_date: Any) -> None:
    if not _present(raw_date):
        return
    converted = _gedcom_date(raw_date)
    if converted:
        lines.append(f"{level} DATE {converted}")
    else:
        lines.extend(_wrapped_lines(level, "NOTE", f"Legacy date: {_text(raw_date).strip()}"))


def _append_event(
    lines: list[str],
    tag: str,
    type_name: str,
    date: Any,
    place: str,
    description: Any,
    notes: Any,
    citations: Iterable[tuple[_Row, str]],
    media: Iterable[_Row],
    dataset: str,
) -> None:
    value = _clean_gedcom_text(description).replace("\n", " ").strip()
    if value and len(value.encode("utf-8")) <= 180:
        lines.append(f"1 {tag} {value}")
    else:
        lines.append(f"1 {tag}")
        if value:
            lines.extend(_wrapped_lines(2, "NOTE", description))
    if tag == "EVEN":
        lines.extend(_wrapped_lines(2, "TYPE", type_name))
    _append_date(lines, 2, date)
    if place:
        lines.extend(_wrapped_lines(2, "PLAC", place))
    if _present(notes):
        lines.extend(_wrapped_lines(2, "NOTE", notes))
    for citation, source_xref in citations:
        for line in _citation_lines(citation, source_xref):
            current_level, remainder = line.split(" ", 1)
            lines.append(f"{int(current_level) + 1} {remainder}")
    for medium in media:
        lines.append(f"2 OBJE {_xref('M', dataset, _record_id(medium, 'media'))}")


def _family_partners(row: Mapping[str, Any]) -> tuple[Any, Any]:
    first = _get(
        row,
        "husband_id",
        "husband_individual_id",
        "spouse1_id",
        "partner1_id",
        "person1_id",
        "father_id",
    )
    second = _get(
        row,
        "wife_id",
        "wife_individual_id",
        "spouse2_id",
        "partner2_id",
        "person2_id",
        "mother_id",
    )
    return first, second


def _child_link(row: Mapping[str, Any]) -> tuple[Any, Any]:
    return (
        _get(row, "family_id", "marriage_id", "parent_family_id"),
        _get(row, "individual_id", "person_id", "child_individual_id", "child_id"),
    )


def export_gedcom(
    db_path: str | Path,
    output_path: str | Path,
    dataset_id: int | str | None = None,
    *,
    include_private: bool = True,
) -> None:
    """Export a merged SQLite database as UTF-8 GEDCOM 5.5.1.

    All source data, including private and living records, is exported by
    default for archival use.  Set ``include_private=False`` to omit people
    marked private or living and links or records owned by those people.
    """
    with _connect_read_only(db_path) as connection:
        schema = _Schema(connection)
        people = _sort_rows(schema.rows("people", dataset_id), dataset_id, "person")
        alternate_names = schema.rows("alternate_names", dataset_id)
        families = _sort_rows(schema.rows("families", dataset_id), dataset_id, "family")
        children = schema.rows("children", dataset_id)
        events = _sort_rows(schema.rows("events", dataset_id), dataset_id, "event")
        event_participants = schema.rows("event_participants", dataset_id)
        event_types = schema.rows("event_types", dataset_id)
        sources = _sort_rows(schema.rows("sources", dataset_id), dataset_id, "source")
        citations = _sort_rows(schema.rows("citations", dataset_id), dataset_id, "citation")
        media = _sort_rows(schema.rows("media", dataset_id), dataset_id, "media")
        media_paths = schema.rows("media_paths", dataset_id)
        locations = schema.rows("locations", dataset_id)

    all_people_lookup = _lookup(people, dataset_id, "person")
    all_family_lookup = _lookup(families, dataset_id, "family")
    people_by_legacy_id = _column_lookup(people, dataset_id, "legacy_id")
    families_by_legacy_id = _column_lookup(families, dataset_id, "legacy_id")
    events = _normalise_event_owners(
        events,
        dataset_id,
        people_by_legacy_id,
        families_by_legacy_id,
        all_people_lookup,
        all_family_lookup,
    )
    alternate_name_lookup = _column_lookup(alternate_names, dataset_id, "alternate_name_id", "id")
    child_lookup = _column_lookup(children, dataset_id, "child_id", "id")
    event_participant_lookup = _column_lookup(
        event_participants, dataset_id, "event_participant_id", "id"
    )

    if not include_private:
        people = [row for row in people if not _private_person(row)]

    people_lookup = _lookup(people, dataset_id, "person")
    location_lookup = _lookup(locations, dataset_id, "location")
    source_lookup = _lookup(sources, dataset_id, "source")
    event_type_lookup = _lookup(event_types, dataset_id, "event_type")
    media_path_lookup = _media_path_lookup(media_paths, dataset_id)
    alternate_names_by_person: dict[_Key, list[_Row]] = defaultdict(list)
    for alternate_name in alternate_names:
        alternate_names_by_person[
            _key(
                _dataset(alternate_name, dataset_id),
                _get(alternate_name, "individual_id", "person_id"),
            )
        ].append(alternate_name)

    def included_person(dataset: str, person_id: Any) -> bool:
        return not _has_id(person_id) or _resolve_row(people_lookup, dataset, person_id) is not None

    kept_families: list[_Row] = []
    for family in families:
        family_dataset = _dataset(family, dataset_id)
        first, second = _family_partners(family)
        if not include_private and _truthy(_get(family, "private", "is_private")):
            continue
        if included_person(family_dataset, first) and included_person(family_dataset, second):
            kept_families.append(family)
    families = kept_families
    family_lookup = _lookup(families, dataset_id, "family")

    children_by_family: dict[_Key, list[Any]] = defaultdict(list)
    parent_families_by_child: dict[_Key, list[Any]] = defaultdict(list)
    for child_row in children:
        child_dataset = _dataset(child_row, dataset_id)
        family_id, child_id = _child_link(child_row)
        if (
            _resolve_row(family_lookup, child_dataset, family_id) is None
            or _resolve_row(people_lookup, child_dataset, child_id) is None
        ):
            continue
        children_by_family[_key(child_dataset, family_id)].append(child_id)
        parent_families_by_child[_key(child_dataset, child_id)].append(family_id)

    spouse_families_by_person: dict[_Key, list[Any]] = defaultdict(list)
    for family in families:
        family_dataset = _dataset(family, dataset_id)
        family_id = _record_id(family, "family")
        for partner_id in _family_partners(family):
            if _has_id(partner_id):
                spouse_families_by_person[_key(family_dataset, partner_id)].append(family_id)

    events = [
        event
        for event in events
        if (
            (
                _owner(event)[0] == "person"
                and included_person(_dataset(event, dataset_id), _owner(event)[1])
            )
            or (
                _owner(event)[0] == "family"
                and _resolve_row(family_lookup, _dataset(event, dataset_id), _owner(event)[1])
                is not None
            )
            or _owner(event)[0] not in {"person", "family"}
        )
        and (include_private or not _truthy(_get(event, "private", "is_private")))
    ]
    event_lookup = _lookup(events, dataset_id, "event")

    person_target_kinds = {"person", "birth", "christening", "death", "burial"}

    def target_is_included(dataset: str, kind: str, owner_id: Any) -> bool:
        if kind in person_target_kinds:
            return _resolve_row(people_lookup, dataset, owner_id) is not None
        if kind == "family":
            return _resolve_row(family_lookup, dataset, owner_id) is not None
        if kind == "event":
            return _resolve_row(event_lookup, dataset, owner_id) is not None
        return True

    citations = [
        citation
        for citation in citations
        if any(
            target_is_included(_dataset(citation, dataset_id), kind, owner_id)
            for kind, owner_id in _citation_targets(
                citation,
                alternate_name_lookup,
                child_lookup,
                event_participant_lookup,
                dataset_id,
            )
        )
    ]
    citation_lookup = _lookup(citations, dataset_id, "citation")
    citations_by_owner = _citation_map(
        citations,
        source_lookup,
        alternate_name_lookup,
        child_lookup,
        event_participant_lookup,
        dataset_id,
    )

    def medium_is_included(medium: Mapping[str, Any]) -> bool:
        dataset = _dataset(medium, dataset_id)
        kind, owner_id = _media_owner(medium)
        if kind == "citation":
            return _resolve_row(citation_lookup, dataset, owner_id) is not None
        return target_is_included(dataset, kind, owner_id)

    media = [
        medium
        for medium in media
        if medium_is_included(medium)
        and _has_id(_record_id(medium, "media"))
        and _present(_media_file(medium, _dataset(medium, dataset_id), media_path_lookup))
    ]

    events_by_owner: dict[tuple[str, str, str], list[_Row]] = defaultdict(list)
    for event in events:
        event_dataset = _dataset(event, dataset_id)
        kind, owner_id = _owner(event)
        if kind == "person" and _resolve_row(people_lookup, event_dataset, owner_id) is None:
            continue
        if kind == "family" and _resolve_row(family_lookup, event_dataset, owner_id) is None:
            continue
        events_by_owner[(event_dataset, kind, _id(owner_id))].append(event)

    media_by_owner: dict[tuple[str, str, str], list[_Row]] = defaultdict(list)
    for medium in media:
        medium_dataset = _dataset(medium, dataset_id)
        kind, owner_id = _media_owner(medium)
        media_by_owner[(medium_dataset, kind, _id(owner_id))].append(medium)

    lines = [
        "0 HEAD",
        "1 SOUR LegacyFamilyTreeReader",
        "2 VERS 0.3.0",
        "2 NAME Legacy Family Tree Reader",
        "1 GEDC",
        "2 VERS 5.5.1",
        "2 FORM LINEAGE-LINKED",
        "1 CHAR UTF-8",
    ]

    for person in people:
        dataset = _dataset(person, dataset_id)
        person_id = _record_id(person, "person")
        lines.append(f"0 {_xref('I', dataset, person_id)} INDI")
        lines.extend(_wrapped_lines(1, "NAME", _gedcom_name(person)))
        given = _get(person, "given_name", "given", "first_name")
        surname = _get(person, "surname", "last_name", "family_name")
        prefix = _get(person, "prefix", "name_prefix")
        suffix = _get(person, "suffix", "name_suffix", "title", "title_suffix")
        if _present(given):
            lines.extend(_wrapped_lines(2, "GIVN", given))
        if _present(surname):
            lines.extend(_wrapped_lines(2, "SURN", surname))
        if _present(prefix):
            lines.extend(_wrapped_lines(2, "NPFX", prefix))
        if _present(suffix):
            lines.extend(_wrapped_lines(2, "NSFX", suffix))
        for alternate_name in alternate_names_by_person.get(_key(dataset, person_id), []):
            lines.extend(_wrapped_lines(1, "NAME", _gedcom_name(alternate_name)))
            lines.append("2 TYPE aka")
            alternate_given = _get(alternate_name, "given_name", "given_names")
            alternate_surname = _get(alternate_name, "surname")
            if _present(alternate_given):
                lines.extend(_wrapped_lines(2, "GIVN", alternate_given))
            if _present(alternate_surname):
                lines.extend(_wrapped_lines(2, "SURN", alternate_surname))
            alternate_note = _get(alternate_name, "alternate_name_note", "notes", "note")
            if _present(alternate_note):
                lines.extend(_wrapped_lines(2, "NOTE", alternate_note))
        lines.append(f"1 SEX {_person_sex(person)}")

        for fact_name, tag, date_names, location_id_names, place_names, note_names in _VITAL_FACTS:
            raw_date = _get(person, *date_names)
            fact_place = _place(person, dataset, location_lookup, location_id_names, place_names)
            fact_notes = _get(person, *note_names)
            fact_kind = fact_name.casefold()
            fact_citations = citations_by_owner.get((dataset, fact_kind, _id(person_id)), ())
            fact_media = media_by_owner.get((dataset, fact_kind, _id(person_id)), ())
            if not any(
                (
                    _present(raw_date),
                    bool(fact_place),
                    _present(fact_notes),
                    bool(fact_citations),
                    bool(fact_media),
                )
            ):
                continue
            lines.append(f"1 {tag}")
            _append_date(lines, 2, raw_date)
            if fact_place:
                lines.extend(_wrapped_lines(2, "PLAC", fact_place))
            if fact_name == "Death":
                cause = _get(person, "death_cause", "cause_of_death")
                if _present(cause):
                    lines.extend(_wrapped_lines(2, "CAUS", cause))
            if _present(fact_notes):
                lines.extend(_wrapped_lines(2, "NOTE", fact_notes))
            _append_citations(
                lines,
                citations_by_owner,
                dataset,
                (fact_kind, tag.casefold()),
                person_id,
                level=2,
            )
            for medium in fact_media:
                lines.append(f"2 OBJE {_xref('M', dataset, _record_id(medium, 'media'))}")

        if _truthy(_get(person, "cremated", "is_cremated")):
            lines.append("1 CREM")
        for note_names, label in (
            (("name_notes", "name_note"), "Name note"),
            (("notes", "general_notes"), ""),
            (("references", "research_notes"), "References"),
            (("medical_notes", "medical"), "Medical"),
        ):
            value = _get(person, *note_names)
            if _present(value):
                note = f"{label}: {_text(value)}" if label else value
                lines.extend(_wrapped_lines(1, "NOTE", note))

        for event in events_by_owner.get((dataset, "person", _id(person_id)), []):
            event_id = _record_id(event, "event")
            type_name, tag = _event_type(event, dataset, event_type_lookup)
            event_place = _place(
                event,
                dataset,
                location_lookup,
                ("event_location_id", "location_id", "place_id"),
                ("location", "place"),
            )
            _append_event(
                lines,
                tag,
                type_name,
                _get(event, "event_date", "date"),
                event_place,
                _get(event, "description", "event_description", "value"),
                _get(event, "notes", "note"),
                citations_by_owner.get((dataset, "event", _id(event_id)), ()),
                media_by_owner.get((dataset, "event", _id(event_id)), ()),
                dataset,
            )

        for family_id in parent_families_by_child.get(_key(dataset, person_id), []):
            lines.append(f"1 FAMC {_xref('F', dataset, family_id)}")
        for family_id in spouse_families_by_person.get(_key(dataset, person_id), []):
            lines.append(f"1 FAMS {_xref('F', dataset, family_id)}")
        for medium in media_by_owner.get((dataset, "person", _id(person_id)), []):
            lines.append(f"1 OBJE {_xref('M', dataset, _record_id(medium, 'media'))}")
        _append_citations(lines, citations_by_owner, dataset, ("person", "individual"), person_id)

    for family in families:
        dataset = _dataset(family, dataset_id)
        family_id = _record_id(family, "family")
        first, second = _family_partners(family)
        lines.append(f"0 {_xref('F', dataset, family_id)} FAM")
        if _resolve_row(people_lookup, dataset, first):
            lines.append(f"1 HUSB {_xref('I', dataset, first)}")
        if _resolve_row(people_lookup, dataset, second):
            lines.append(f"1 WIFE {_xref('I', dataset, second)}")

        marriage_date = _get(family, "marriage_date", "married_date", "date")
        marriage_place = _place(
            family,
            dataset,
            location_lookup,
            ("marriage_location_id", "marriage_place_id", "location_id", "place_id"),
            ("marriage_location", "marriage_place", "location", "place"),
        )
        marriage_notes = _get(family, "marriage_notes", "notes", "note")
        if any((_present(marriage_date), bool(marriage_place), _present(marriage_notes))):
            lines.append("1 MARR")
            _append_date(lines, 2, marriage_date)
            if marriage_place:
                lines.extend(_wrapped_lines(2, "PLAC", marriage_place))
            if _present(marriage_notes):
                lines.extend(_wrapped_lines(2, "NOTE", marriage_notes))
            _append_citations(
                lines,
                citations_by_owner,
                dataset,
                ("marriage", "marr"),
                family_id,
                level=2,
            )
        end_date = _get(family, "marriage_end_date", "end_date", "divorce_date")
        if _present(end_date):
            lines.append("1 DIV")
            _append_date(lines, 2, end_date)
        for child_id in children_by_family.get(_key(dataset, family_id), []):
            lines.append(f"1 CHIL {_xref('I', dataset, child_id)}")
        for event in events_by_owner.get((dataset, "family", _id(family_id)), []):
            event_id = _record_id(event, "event")
            type_name, tag = _event_type(event, dataset, event_type_lookup)
            event_place = _place(
                event,
                dataset,
                location_lookup,
                ("event_location_id", "location_id", "place_id"),
                ("location", "place"),
            )
            _append_event(
                lines,
                tag,
                type_name,
                _get(event, "event_date", "date"),
                event_place,
                _get(event, "description", "event_description", "value"),
                _get(event, "notes", "note"),
                citations_by_owner.get((dataset, "event", _id(event_id)), ()),
                media_by_owner.get((dataset, "event", _id(event_id)), ()),
                dataset,
            )
        for medium in media_by_owner.get((dataset, "family", _id(family_id)), []):
            lines.append(f"1 OBJE {_xref('M', dataset, _record_id(medium, 'media'))}")
        _append_citations(lines, citations_by_owner, dataset, ("family", "marriage"), family_id)

    for source in sources:
        dataset = _dataset(source, dataset_id)
        source_id = _record_id(source, "source")
        lines.append(f"0 {_xref('S', dataset, source_id)} SOUR")
        for tag, names in (
            ("ABBR", ("name", "source_name", "abbreviation")),
            ("TITL", ("title", "source_title")),
            ("AUTH", ("author", "source_author")),
            ("PUBL", ("publication", "publisher", "source_publication")),
            ("TEXT", ("text", "source_text", "contents")),
            ("NOTE", ("notes", "note", "source_notes", "source_note")),
        ):
            value = _get(source, *names)
            if _present(value):
                lines.extend(_wrapped_lines(1, tag, value))
        call_number = _get(source, "call_number", "call_num", "source_call_num")
        if _present(call_number):
            lines.extend(_wrapped_lines(1, "NOTE", f"Call number: {_text(call_number)}"))
        url = _get(source, "url", "web_address")
        if _present(url):
            lines.extend(_wrapped_lines(1, "NOTE", f"URL: {_text(url)}"))
        for medium in media_by_owner.get((dataset, "source", _id(source_id)), []):
            lines.append(f"1 OBJE {_xref('M', dataset, _record_id(medium, 'media'))}")

    for medium in media:
        dataset = _dataset(medium, dataset_id)
        media_id = _record_id(medium, "media")
        path = _media_file(medium, dataset, media_path_lookup)
        if not _has_id(media_id) or not _present(path):
            continue
        lines.append(f"0 {_xref('M', dataset, media_id)} OBJE")
        lines.extend(_wrapped_lines(1, "FILE", path))
        suffix = Path(_text(path)).suffix.removeprefix(".").lower()
        if suffix:
            lines.append(f"2 FORM {suffix}")
        title = _get(medium, "caption", "media_caption", "title", "name")
        if _present(title):
            lines.extend(_wrapped_lines(2, "TITL", title))
        notes = _get(medium, "description", "media_desc", "notes", "note")
        if _present(notes):
            lines.extend(_wrapped_lines(1, "NOTE", notes))

    lines.extend(("0 TRLR", ""))
    Path(output_path).write_text("\r\n".join(lines), encoding="utf-8", newline="")


def _display_person(people_lookup: Mapping[_Key, _Row], dataset: Any, person_id: Any) -> str:
    person = _resolve_row(people_lookup, dataset, person_id)
    if person is None:
        return ""
    name = _full_name(person)
    return f"{name} [{_id(person_id)}]" if name else _id(person_id)


def _display_family(
    family: Mapping[str, Any], people_lookup: Mapping[_Key, _Row], dataset: Any
) -> str:
    first, second = _family_partners(family)
    names = [
        _display_person(people_lookup, dataset, person_id)
        for person_id in (first, second)
        if _has_id(person_id)
    ]
    return " + ".join(name for name in names if name)


def _excel_value(value: Any) -> Any:
    if isinstance(value, bytes):
        value = value.decode("utf-8", errors="replace")
    if isinstance(value, str):
        return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)
    if isinstance(value, (int, float, bool)) or value is None:
        return value
    return str(value)


def _write_sheet(
    workbook: Workbook, title: str, headers: Sequence[str], rows: Iterable[Sequence[Any]]
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(list(headers))
    for row in rows:
        sheet.append([_excel_value(value) for value in row])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for column_number, header in enumerate(headers, start=1):
        values = (
            sheet.cell(row=row, column=column_number).value for row in range(1, sheet.max_row + 1)
        )
        width = max(len(str(value)) for value in values if value is not None)
        sheet.column_dimensions[get_column_letter(column_number)].width = min(
            max(width + 2, 10), 50
        )


def export_excel(
    db_path: str | Path,
    output_path: str | Path,
    dataset_id: int | str | None = None,
    *,
    include_private: bool = True,
) -> None:
    """Export query-friendly worksheets from a merged SQLite database.

    All source data, including private and living records, is included by
    default.  Set ``include_private=False`` to omit people marked private or
    living and dependent family, event, citation, and media rows.
    """
    with _connect_read_only(db_path) as connection:
        schema = _Schema(connection)
        datasets = _sort_rows(schema.rows("datasets", dataset_id), dataset_id, "dataset")
        people = _sort_rows(schema.rows("people", dataset_id), dataset_id, "person")
        alternate_names = schema.rows("alternate_names", dataset_id)
        families = _sort_rows(schema.rows("families", dataset_id), dataset_id, "family")
        children = schema.rows("children", dataset_id)
        events = _sort_rows(schema.rows("events", dataset_id), dataset_id, "event")
        event_participants = schema.rows("event_participants", dataset_id)
        event_types = schema.rows("event_types", dataset_id)
        sources = _sort_rows(schema.rows("sources", dataset_id), dataset_id, "source")
        source_types = schema.rows("source_types", dataset_id)
        citations = _sort_rows(schema.rows("citations", dataset_id), dataset_id, "citation")
        media = _sort_rows(schema.rows("media", dataset_id), dataset_id, "media")
        media_paths = schema.rows("media_paths", dataset_id)
        locations = _sort_rows(schema.rows("locations", dataset_id), dataset_id, "location")
        marriage_statuses = schema.rows("marriage_statuses", dataset_id)
        child_statuses = schema.rows("child_statuses", dataset_id)
        child_relationships = schema.rows("child_relationships", dataset_id)

    all_people_lookup = _lookup(people, dataset_id, "person")
    all_family_lookup = _lookup(families, dataset_id, "family")
    events = _normalise_event_owners(
        events,
        dataset_id,
        _column_lookup(people, dataset_id, "legacy_id"),
        _column_lookup(families, dataset_id, "legacy_id"),
        all_people_lookup,
        all_family_lookup,
    )
    alternate_name_lookup = _column_lookup(alternate_names, dataset_id, "alternate_name_id", "id")
    child_lookup = _column_lookup(children, dataset_id, "child_id", "id")
    event_participant_lookup = _column_lookup(
        event_participants, dataset_id, "event_participant_id", "id"
    )

    if not include_private:
        people = [row for row in people if not _private_person(row)]
    people_lookup = _lookup(people, dataset_id, "person")
    location_lookup = _lookup(locations, dataset_id, "location")
    family_lookup_all = _lookup(families, dataset_id, "family")
    source_lookup = _lookup(sources, dataset_id, "source")
    event_lookup = _lookup(events, dataset_id, "event")
    event_type_lookup = _lookup(event_types, dataset_id, "event_type")
    media_path_lookup = _media_path_lookup(media_paths, dataset_id)

    def person_is_included(dataset: str, person_id: Any) -> bool:
        return not _has_id(person_id) or _resolve_row(people_lookup, dataset, person_id) is not None

    families = [
        family
        for family in families
        if not (not include_private and _truthy(_get(family, "private", "is_private")))
        and all(
            person_is_included(_dataset(family, dataset_id), partner)
            for partner in _family_partners(family)
        )
    ]
    family_lookup = _lookup(families, dataset_id, "family")

    children = [
        row
        for row in children
        if _resolve_row(family_lookup, _dataset(row, dataset_id), _child_link(row)[0])
        and _resolve_row(people_lookup, _dataset(row, dataset_id), _child_link(row)[1])
    ]
    events = [
        row
        for row in events
        if (
            (
                _owner(row)[0] == "person"
                and _resolve_row(people_lookup, _dataset(row, dataset_id), _owner(row)[1])
            )
            or (
                _owner(row)[0] == "family"
                and _resolve_row(family_lookup, _dataset(row, dataset_id), _owner(row)[1])
            )
            or _owner(row)[0] not in {"person", "family"}
        )
        and (include_private or not _truthy(_get(row, "private", "is_private")))
    ]
    event_lookup = _lookup(events, dataset_id, "event")

    person_target_kinds = {"person", "birth", "christening", "death", "burial"}

    def target_is_included(dataset: str, kind: str, owner_id: Any) -> bool:
        if kind in person_target_kinds:
            return _resolve_row(people_lookup, dataset, owner_id) is not None
        if kind == "family":
            return _resolve_row(family_lookup, dataset, owner_id) is not None
        if kind == "event":
            return _resolve_row(event_lookup, dataset, owner_id) is not None
        return True

    citations = [
        row
        for row in citations
        if any(
            target_is_included(_dataset(row, dataset_id), kind, owner_id)
            for kind, owner_id in _citation_targets(
                row,
                alternate_name_lookup,
                child_lookup,
                event_participant_lookup,
                dataset_id,
            )
        )
    ]
    citation_lookup = _lookup(citations, dataset_id, "citation")

    def medium_is_included(row: Mapping[str, Any]) -> bool:
        dataset = _dataset(row, dataset_id)
        kind, owner_id = _media_owner(row)
        if kind == "citation":
            return _resolve_row(citation_lookup, dataset, owner_id) is not None
        return target_is_included(dataset, kind, owner_id)

    media = [row for row in media if medium_is_included(row)]

    status_lookup: dict[_Key, _Row] = {}
    for row in marriage_statuses:
        status_id = _get(row, "marriage_status_id", "status_id", "id")
        if _has_id(status_id):
            status_lookup[_key(_dataset(row, dataset_id), status_id)] = row
    child_status_lookup: dict[_Key, _Row] = {}
    for row in child_statuses:
        status_id = _get(row, "child_status_id", "status_id", "id")
        if _has_id(status_id):
            child_status_lookup[_key(_dataset(row, dataset_id), status_id)] = row
    relationship_lookup: dict[_Key, _Row] = {}
    for row in child_relationships:
        relationship_id = _get(
            row,
            "relationship_id",
            "child_relationship_id",
            "child_relationship_type_id",
            "id",
        )
        if _has_id(relationship_id):
            relationship_lookup[_key(_dataset(row, dataset_id), relationship_id)] = row
    source_type_lookup: dict[_Key, _Row] = {}
    for row in source_types:
        source_type_id = _get(row, "source_type_id", "type_id", "id")
        if _has_id(source_type_id):
            source_type_lookup[_key(_dataset(row, dataset_id), source_type_id)] = row

    workbook = Workbook()
    default_sheet: Worksheet = workbook.active
    workbook.remove(default_sheet)

    if not datasets:
        seen_datasets = sorted({_dataset(row, dataset_id) for row in people + families})
        datasets = [{"datasetid": value} for value in seen_datasets]
    _write_sheet(
        workbook,
        "Data Sets",
        ("Dataset ID", "Name", "Source Path", "Imported At", "Description"),
        (
            (
                _record_id(row, "dataset"),
                _get(row, "name", "dataset_name", "source_name"),
                _get(row, "source_path", "path", "file_path"),
                _get(row, "imported_at", "created_at", "import_date"),
                _get(row, "description", "notes"),
            )
            for row in datasets
        ),
    )

    people_headers = (
        "Dataset ID",
        "Person ID",
        "Full Name",
        "Given Name",
        "Surname",
        "Prefix",
        "Suffix/Title",
        "Sex",
        "Living",
        "Private",
        "Birth Date",
        "Birth Place",
        "Death Date",
        "Death Place",
        "User Reference",
    )
    _write_sheet(
        workbook,
        "People",
        people_headers,
        (
            (
                (dataset := _dataset(row, dataset_id)),
                _record_id(row, "person"),
                _full_name(row),
                _get(row, "given_name", "given_names", "given", "first_name"),
                _get(row, "surname", "last_name", "family_name"),
                _get(row, "prefix", "name_prefix", "title_prefix"),
                _get(row, "suffix", "name_suffix", "title", "title_suffix"),
                _person_sex(row),
                _truthy(_get(row, "living", "is_living", "living_flag")),
                _truthy(_get(row, "private", "is_private", "private_flag")),
                _get(row, "birth_date", "birth_d"),
                _place(
                    row,
                    dataset,
                    location_lookup,
                    ("birth_location_id", "birth_place_id"),
                    ("birth_location", "birth_place"),
                ),
                _get(row, "death_date", "death_d"),
                _place(
                    row,
                    dataset,
                    location_lookup,
                    ("death_location_id", "death_place_id"),
                    ("death_location", "death_place"),
                ),
                _get(
                    row,
                    "user_reference",
                    "user_ref",
                    "reference_number",
                    "legacy_rin",
                    "legacy_id",
                ),
            )
            for row in people
        ),
    )

    def family_status(row: Mapping[str, Any], dataset: str) -> str:
        explicit = _get(row, "status", "marriage_status", "status_name")
        if _present(explicit):
            return _text(explicit)
        status = _resolve_row(status_lookup, dataset, _get(row, "marriage_status_id", "status_id"))
        return _text(_get(status or {}, "status", "marriage_status", "name"))

    _write_sheet(
        workbook,
        "Families",
        (
            "Dataset ID",
            "Family ID",
            "Family",
            "Partner 1 ID",
            "Partner 1",
            "Partner 2 ID",
            "Partner 2",
            "Marriage Date",
            "Marriage Place",
            "End Date",
            "Status",
            "Private",
            "Notes",
        ),
        (
            (
                (dataset := _dataset(row, dataset_id)),
                _record_id(row, "family"),
                _display_family(row, people_lookup, dataset),
                (partners := _family_partners(row))[0],
                _display_person(people_lookup, dataset, partners[0]),
                partners[1],
                _display_person(people_lookup, dataset, partners[1]),
                _get(row, "marriage_date", "married_date", "date"),
                _place(
                    row,
                    dataset,
                    location_lookup,
                    ("marriage_location_id", "marriage_place_id", "location_id", "place_id"),
                    ("marriage_location", "marriage_place", "location", "place"),
                ),
                _get(row, "marriage_end_date", "end_date", "divorce_date"),
                family_status(row, dataset),
                _truthy(_get(row, "private", "is_private")),
                _get(row, "notes", "note", "marriage_notes"),
            )
            for row in families
        ),
    )

    def child_label(
        lookup: Mapping[_Key, _Row], dataset: str, value: Any, names: Sequence[str]
    ) -> str:
        row = _resolve_row(lookup, dataset, value)
        return _text(_get(row or {}, *names))

    _write_sheet(
        workbook,
        "Parent Child",
        (
            "Dataset ID",
            "Family ID",
            "Family",
            "Child ID",
            "Child",
            "Child Order",
            "Child Status",
            "Parent 1 Relationship",
            "Parent 2 Relationship",
            "Notes",
        ),
        (
            (
                (dataset := _dataset(row, dataset_id)),
                (link := _child_link(row))[0],
                _display_family(
                    _resolve_row(family_lookup_all, dataset, link[0]) or {},
                    people_lookup,
                    dataset,
                ),
                link[1],
                _display_person(people_lookup, dataset, link[1]),
                _get(row, "child_order", "display_order", "order", "sort_order"),
                child_label(
                    child_status_lookup,
                    dataset,
                    _get(row, "child_status_id", "status_id"),
                    ("child_status", "status", "name"),
                )
                or _text(_get(row, "child_status", "status")),
                child_label(
                    relationship_lookup,
                    dataset,
                    _get(
                        row,
                        "parent1_relationship_id",
                        "father_relationship_id",
                        "father_relationship_type_id",
                    ),
                    ("relationship", "child_parent_relation", "name"),
                ),
                child_label(
                    relationship_lookup,
                    dataset,
                    _get(
                        row,
                        "parent2_relationship_id",
                        "mother_relationship_id",
                        "mother_relationship_type_id",
                    ),
                    ("relationship", "child_parent_relation", "name"),
                ),
                _get(row, "notes", "note"),
            )
            for row in children
        ),
    )

    fact_rows: list[tuple[Any, ...]] = []
    for person in people:
        dataset = _dataset(person, dataset_id)
        person_id = _record_id(person, "person")
        owner_name = _full_name(person)
        for fact_name, _tag, date_names, location_id_names, place_names, note_names in _VITAL_FACTS:
            date = _get(person, *date_names)
            place = _place(person, dataset, location_lookup, location_id_names, place_names)
            notes = _get(person, *note_names)
            if any((_present(date), bool(place), _present(notes))):
                fact_rows.append(
                    (dataset, "Person", person_id, owner_name, fact_name, date, place, "", notes)
                )
        for fact_name, names in (
            ("General Notes", ("notes", "note")),
            ("Name Notes", ("name_notes", "name_note")),
            ("References", ("references",)),
            ("Medical", ("medical_notes", "medical")),
            ("DNA", ("dna_notes", "dna")),
        ):
            notes = _get(person, *names)
            if _present(notes):
                fact_rows.append(
                    (dataset, "Person", person_id, owner_name, fact_name, "", "", "", notes)
                )
    for family in families:
        dataset = _dataset(family, dataset_id)
        family_id = _record_id(family, "family")
        family_name = _display_family(family, people_lookup, dataset)
        date = _get(family, "marriage_date", "married_date", "date")
        place = _place(
            family,
            dataset,
            location_lookup,
            ("marriage_location_id", "marriage_place_id", "location_id", "place_id"),
            ("marriage_location", "marriage_place", "location", "place"),
        )
        notes = _get(family, "notes", "note", "marriage_notes")
        if any((_present(date), bool(place), _present(notes))):
            fact_rows.append(
                (dataset, "Family", family_id, family_name, "Marriage", date, place, "", notes)
            )
    _write_sheet(
        workbook,
        "Facts and Notes",
        (
            "Dataset ID",
            "Owner Type",
            "Owner ID",
            "Owner Name",
            "Fact Type",
            "Date",
            "Place",
            "Value",
            "Notes",
        ),
        fact_rows,
    )

    def owner_name(dataset: str, kind: str, owner_id: Any) -> str:
        if kind in person_target_kinds:
            return _display_person(people_lookup, dataset, owner_id)
        if kind == "family":
            family = _resolve_row(family_lookup, dataset, owner_id)
            return _display_family(family or {}, people_lookup, dataset)
        if kind == "event":
            event = _resolve_row(event_lookup, dataset, owner_id)
            return _event_type(event or {}, dataset, event_type_lookup)[0] if event else ""
        return ""

    _write_sheet(
        workbook,
        "Events",
        (
            "Dataset ID",
            "Event ID",
            "Owner Type",
            "Owner ID",
            "Owner Name",
            "Event Type",
            "GEDCOM Tag",
            "Date",
            "Place",
            "Description",
            "Notes",
            "Private",
        ),
        (
            (
                (dataset := _dataset(row, dataset_id)),
                _record_id(row, "event"),
                (owner := _owner(row))[0].title(),
                owner[1],
                owner_name(dataset, owner[0], owner[1]),
                (event_type := _event_type(row, dataset, event_type_lookup))[0],
                event_type[1],
                _get(row, "event_date", "date"),
                _place(
                    row,
                    dataset,
                    location_lookup,
                    ("event_location_id", "location_id", "place_id"),
                    ("location", "place"),
                ),
                _get(row, "description", "event_description", "value"),
                _get(row, "notes", "note"),
                _truthy(_get(row, "private", "is_private")),
            )
            for row in events
        ),
    )

    _write_sheet(
        workbook,
        "Sources",
        (
            "Dataset ID",
            "Source ID",
            "Name",
            "Title",
            "Author",
            "Publication",
            "Source Type",
            "Call Number",
            "URL",
            "Text",
            "Notes",
        ),
        (
            (
                _dataset(row, dataset_id),
                _record_id(row, "source"),
                _get(row, "name", "source_name", "abbreviation"),
                _get(row, "title", "source_title"),
                _get(row, "author", "source_author"),
                _get(row, "publication", "publisher", "source_publication"),
                _get(
                    _resolve_row(
                        source_type_lookup,
                        _dataset(row, dataset_id),
                        _get(row, "source_type_id", "type_id"),
                    )
                    or row,
                    "source_type",
                    "type",
                    "name",
                ),
                _get(row, "call_number", "call_num", "source_call_num"),
                _get(row, "url", "web_address"),
                _get(row, "text", "source_text", "contents"),
                _get(row, "notes", "note", "source_notes", "source_note"),
            )
            for row in sources
        ),
    )

    _write_sheet(
        workbook,
        "Citations",
        (
            "Dataset ID",
            "Citation ID",
            "Owner Type",
            "Owner ID",
            "Owner Name",
            "Source ID",
            "Source Name",
            "Detail/Page",
            "Text",
            "Notes",
            "Quality",
        ),
        (
            (
                (dataset := _dataset(row, dataset_id)),
                _record_id(row, "citation"),
                (
                    owner := _citation_target(
                        row,
                        alternate_name_lookup,
                        child_lookup,
                        event_participant_lookup,
                        dataset_id,
                    )
                )[0].title(),
                owner[1],
                owner_name(dataset, owner[0], owner[1]),
                (source_id := _get(row, "source_id")),
                _get(
                    _resolve_row(source_lookup, dataset, source_id) or {},
                    "name",
                    "source_name",
                    "title",
                    "source_title",
                ),
                _get(row, "detail", "source_detail", "page", "where_within_source"),
                _get(
                    row,
                    "text",
                    "source_text",
                    "detail_text",
                    "source_detail_text",
                    "content",
                    "quote",
                ),
                _get(row, "notes", "note", "detail_note", "source_detail_note"),
                _get(row, "quality", "surety", "source_surety", "certainty"),
            )
            for row in citations
        ),
    )

    _write_sheet(
        workbook,
        "Media",
        (
            "Dataset ID",
            "Media ID",
            "Owner Type",
            "Owner ID",
            "Owner Name",
            "Path/URL",
            "Caption",
            "Date",
            "Description",
            "Preferred",
            "Private",
        ),
        (
            (
                _dataset(row, dataset_id),
                _record_id(row, "media"),
                (owner := _media_owner(row))[0].title(),
                owner[1],
                owner_name(_dataset(row, dataset_id), owner[0], owner[1]),
                _media_file(row, _dataset(row, dataset_id), media_path_lookup),
                _get(row, "caption", "media_caption", "title", "name"),
                _get(row, "date", "media_date"),
                _get(row, "description", "media_desc", "notes", "note"),
                _truthy(_get(row, "preferred", "is_preferred", "media_preferred")),
                _truthy(_get(row, "private", "is_private")),
            )
            for row in media
        ),
    )

    _write_sheet(
        workbook,
        "Locations",
        (
            "Dataset ID",
            "Location ID",
            "Location",
            "Short Name",
            "Latitude",
            "Longitude",
            "FamilySearch Place ID",
            "Verified",
            "Notes",
        ),
        (
            (
                _dataset(row, dataset_id),
                _record_id(row, "location"),
                _location_name(row),
                _get(row, "short_name"),
                _get(row, "latitude", "lat"),
                _get(row, "longitude", "lon", "lng"),
                _get(row, "familysearch_place_id", "fs_place_id"),
                _truthy(_get(row, "verified", "is_verified")),
                _get(row, "notes", "note"),
            )
            for row in locations
        ),
    )

    workbook.save(Path(output_path))
