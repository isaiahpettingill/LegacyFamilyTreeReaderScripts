from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from legacy_family_tree_reader.exporters import export_excel, export_gedcom


def _gedcom_records(data: bytes) -> list[str]:
    assert b"\r\n" in data
    assert b"\n" not in data.replace(b"\r\n", b"")
    lines = data.decode("utf-8").split("\r\n")
    assert lines[-1] == ""
    return lines[:-1]


def test_gedcom_has_valid_essentials_links_and_private_filter(
    tmp_path: Path, merged_db: Path
) -> None:
    full_path = tmp_path / "full.ged"
    public_path = tmp_path / "public.ged"
    export_gedcom(merged_db, full_path, 1)
    export_gedcom(merged_db, public_path, 1, include_private=False)

    lines = _gedcom_records(full_path.read_bytes())
    assert lines[0] == "0 HEAD"
    assert lines[-1] == "0 TRLR"
    assert "1 CHAR UTF-8" in lines
    assert "2 VERS 5.5.1" in lines
    assert all(len(line.encode("utf-8")) <= 255 for line in lines)
    assert all(re.match(r"^\d+ (?:@[^@]+@ )?[A-Z0-9_]+(?: .*)?$", line) for line in lines)

    definitions = {
        match.group(1)
        for line in lines
        if (match := re.match(r"^0 (@[^@]+@) (?:INDI|FAM|SOUR|OBJE)$", line))
    }
    pointers = {
        match.group(1)
        for line in lines
        if (match := re.match(r"^[1-9]\d* (?:FAMC|FAMS|HUSB|WIFE|CHIL|SOUR|OBJE) (@[^@]+@)$", line))
    }
    assert pointers <= definitions
    assert len(definitions) == len(set(definitions))
    assert "0 @ID1R6@ INDI" in lines
    assert "1 CHIL @ID1R5@" in lines
    assert "0 @SD1R50@ SOUR" in lines

    public_lines = _gedcom_records(public_path.read_bytes())
    assert "0 @ID1R6@ INDI" not in public_lines
    assert all("@ID1R6@" not in line for line in public_lines)
    assert "0 @ID1R3@ INDI" in public_lines


def test_excel_export_contains_query_friendly_sheets_and_rows(
    tmp_path: Path, merged_db: Path
) -> None:
    output = tmp_path / "family.xlsx"
    export_excel(merged_db, output, 1)
    workbook = load_workbook(output, read_only=False, data_only=True)
    assert workbook.sheetnames == [
        "Data Sets",
        "People",
        "Families",
        "Parent Child",
        "Facts and Notes",
        "Events",
        "Sources",
        "Citations",
        "Media",
        "Locations",
    ]
    people = workbook["People"]
    assert people.freeze_panes == "A2"
    assert people.auto_filter.ref == people.dimensions
    assert tuple(cell.value for cell in people[1][:5]) == (
        "Dataset ID",
        "Person ID",
        "Full Name",
        "Given Name",
        "Surname",
    )
    assert people.max_row == 7
    assert any(
        row[1] == 3 and row[2] == "Casey Rowan Branch" for row in people.iter_rows(values_only=True)
    )
    assert workbook["Families"].max_row == 3
    assert workbook["Parent Child"].max_row == 4
    assert workbook["Citations"].max_row == 20
